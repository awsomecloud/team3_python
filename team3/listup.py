import feedparser
import openpyxl
from googletrans import Translator

url = "https://aws.amazon.com/about-aws/whats-new/recent/feed/" #what's New list

feed = feedparser.parse(url)

print(len(feed['entries']))

translator = Translator()
result = translator.translate(feed.entries[1].title, dest="ko")
print(result.text)

# print (feed.entries[0].link)
# print (feed.entries[0].title)
# print (feed.entries[0].published)

# xlsx 파일 생성 저장
wb = openpyxl.Workbook()
ws = wb.create_sheet('list')
ws = wb.active
ws['A1'] = 'link'
ws['B1'] = 'date'

for i in range(len(feed.entries)):
    # ws.append([feed.entries[i].link,feed.entries[i].title,feed.entries[i].published])
    trans = translator.translate(feed.entries[i].title, dest="ko")
    trans_title = trans.text
    ws.cell(row=(i+2),column=1).value = '=HYPERLINK("{}", "{}")'.format(feed.entries[i].link, trans_title)
    ws.cell((i+2),2,feed.entries[i].published)

wb.save('whatsnew1.xlsx')










